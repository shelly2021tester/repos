import openpyxl
from pytest_f_api.config.testconfig import datafile_path

class ReadWrite:
    caseid="用例编号"
    casemodule="模块"
    casedescription="用例说明"
    requestaddress="请求地址"
    requestmethod="请求方式"
    requestparam="请求参数"
    requestheader="请求头"
    precondition="前置条件"
    isexec="是否执行"
    status="状态码"
    expectresult="期望结果"
    actualresult="实际结果"
    comment="说明"
    title = [caseid, casemodule, casedescription, requestaddress, requestmethod, requestparam, requestheader,
             precondition, isexec, status, expectresult, actualresult]
    def __init__(self,sheet_num):
        self.sheet_num=sheet_num
        self.title=ReadWrite.title

    # excel读
    def read(self):
        self.wk = openpyxl.load_workbook(datafile_path)
        sheets = self.wk.sheetnames
        sheet_name = self.wk[sheets[self.sheet_num - 1]]
        self.wk.active
        col_values=[]
        title_values=[]
        nrows=sheet_name.max_row
        ncols=sheet_name.max_column
        for i in range(2,nrows+1):
            for j in range(1,ncols+1):
                cell_value=sheet_name.cell(i,j).value
                col_values.append(cell_value)
                row_value=dict(zip(self.title, col_values))
            col_values = []
            title_values.append(row_value)
        self.wk.close()
        return title_values


    # excel写操作
    def write(self,caseid,status):
        self.wk = openpyxl.load_workbook(datafile_path)
        sheets = self.wk.sheetnames
        sheet_name = self.wk[sheets[self.sheet_num - 1]]
        self.wk.active
        nrows=sheet_name.max_row
        ncols = sheet_name.max_column
        for i in range(1,nrows+1):
            if sheet_name.cell(i,1).value==caseid:
                sheet_name.cell(i,ncols-1).value=status
                break
        self.wk.save(datafile_path)
        self.wk.close()


if __name__=='__main__':
    headers=ReadWrite(1)
    cases=headers.read()
    # for row in cases:
    #     url=row[ReadWrite.requestaddress]
    #     method=row[ReadWrite.requestmethod]
    #     params=row[ReadWrite.requestparam]
    #     r_header=row[ReadWrite.requestheader]
    #     print(url,method,params,r_header)
    print(cases)

